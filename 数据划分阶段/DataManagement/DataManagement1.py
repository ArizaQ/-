import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
cols= ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","","","",""]

def main():
    stage1()
    stage2()
    stage3()
    stage4()
    stage5()
    stage6()
    stage7()
    stage8()
    stage9()
    stage10()
    stage11()
    stage12()
def stage1():
    wbr= Workbook() # 实例化: 创建一个对象
    rsheet= wbr.active # 获取当前活跃的sheet
    rsheet.title= "sheet1"
    rsheet['A1']= "2019.12.8-2020.1.8"

    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i= 2;
    for row in sheet.iter_rows(min_row=3892, max_row=4039, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value))==0:
                pass
            else:
                rsheet['A'+str(i)]= cell.value
                print(cell.value,end=',')
                i+=1
        print()
    wbr.save("result1.xlsx")

def stage2():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['B1'] = "2020.1.8-2020.1.22"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=3750, max_row=3891, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['B' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage3():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['C1'] = "2020.1.23-2020.2.7"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=3059, max_row=3749, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['C' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

def stage4():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['D1'] = "2020.2.10-2020.2.13"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=2847, max_row=3058, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['D' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage5():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['E1'] = "2020.2.14-2020.3.9"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=2138, max_row=2846, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['E' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()

    wbr.save("result1.xlsx")
def stage6():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['F1'] = "2020.3.10-2020.3.25"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=1657, max_row=2137, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['F' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage7():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['G1'] = "2020.3.25-2020.4.15"

    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=1247, max_row=1656, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['G' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet.iter_rows(min_row=544, max_row=646, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['G' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage8():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['H1'] = "2020.4.16-2020.4.30"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=1226, max_row=1246, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['H' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet.iter_rows(min_row=285, max_row=543, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['H' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage9():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['I1'] = "2020.5.1-2020.5.15"

    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=970, max_row=1225, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['I' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage10():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['J1'] = "2020.5.16-2020.5.31"

    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=712, max_row=969, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['J' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()

    wbr.save("result1.xlsx")
def stage11():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['K1'] = "2020.6.1-2020.6.15"

    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=647, max_row=711, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['K' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet.iter_rows(min_row=138, max_row=284, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['K' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage12():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['L1'] = "2020.6.16-2020.6.23"
    wb = load_workbook("source.xlsx")
    sheet = wb.get_sheet_by_name("Sheet1")
    i = 2
    for row in sheet.iter_rows(min_row=1, max_row=137, min_col=5, max_col=9):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['L' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

if __name__ == '__main__':
    main()