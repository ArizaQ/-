import datetime

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
cols= ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","","","",""]
'''
一、 创建
'''
# wb= Workbook() # 实例化: 创建一个对象
# sheet= wb.active # 获取当前活跃的sheet
# sheet.title= "张伟旗"
# print(sheet)
# sheet.title= "zhangweiqi"

'''
二、打开已有文件
'''
wb= load_workbook("exceltest1.xlsx")
sheet= wb.active

'''
三、写入数据
'''
# 1.按单元格
# sheet["A5"]= "张"
# sheet["C6"]= "伟旗"
# 2.附加行（下方），从第一列开始
# sheet.append([1,2,3])
# 3.Python 类型自动转换
# sheet["A3"]= datetime.datetime.now().strftime("%Y-%m-%d")
# wb.save("exceltest1.xlsx")
# 保存
# wb.save("exceltest1.xlsx")

'''
四、遍历
'''
wb= load_workbook("NewsWeibo0.xlsx")
# 打印所有sheetnames
# wb.sheetnames   wb.get_sheet_names()
sheet= wb.get_sheet_by_name("疫情新闻")
# 获取单元格： sheet["B5]
# print(sheet["B5"])
# print(sheet["B5"].value)
#
# # 获取多个
# for cell in sheet["B5:B9"]:
#     # cell是一个tuple
#     print(cell[0].value)

# 遍历
# for row in sheet:
#     for cell in row:
#         print(cell.value,end=',')
#     print()
# # row: 一个tuple
#
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value, end=',')
#     print()

# row2-5, col:0-5
for row in sheet.iter_rows(min_row=2, max_row=5, min_col=5, max_col=9):
    for cell in row:
        print(cell.value, end='  ')
    print()

# for column in sheet.iter_cols(min_col= 5, max_col= 9,min_row=2, max_row=5):
#     for cell in column:
#         print(cell.value, end='  ')
#     print()


### 删除：
# wb.remove(sheet)
# del wb[sheet]

'''
五、设置单元格样式
'''
wb= load_workbook("exceltest1.xlsx")
sheet= wb.get_sheet_by_name("张伟旗")
### 字体
# bold_itatic_24_font= Font(name='等线', size=24, itatic=True, color= colors.RED, bold=True)
# sheet["C6"].font= bold_itatic_24_font;
### 对齐方式
sheet["C6"].alignment= Alignment(horizontal='center', vertical='center')
### 长宽高
sheet.row_dimensions[6].height=40
sheet.column_dimensions['C'].width= 30
### 边框/
wb.save("exceltest1.xlsx")
