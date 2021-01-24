import datetime
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
cols= ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","","","",""]
'''
数据时间：
stage3：1.26  2.4
stage4：2.9  2.10  2.12
stage5：2.15  2.18  2.20  2.26  3.2
stage6：3.15  3.18  3.21  3.22  3.23  3.25
stage7：3.26  3.30  4.2  4.4  4.5  4.8  4.10
stage8；4.21  4.22
stage9: 5.2  5.3  5.7  5.15
stage10: 5.18  5.19  5.20  5.23  5.24  
stage11: 6.1  
stage12：6.26  6.27
'''
def main():
    print('hello')
    # stage1()
    # stage2()
    # stage3()
    # stage4()
    # stage5()
    # stage6()
    # stage7()
    # stage8()
    # stage9()
    # stage10()
    # stage11()
    # stage12()
    stage3p()
    stage4p()
    stage5p()
    stage6p()
    stage7p()
    stage8p()
    stage9p()
    stage10p()
    stage11p()
    stage12p()


def stage1():
    wbr= Workbook() # 实例化: 创建一个对象
    rsheet= wbr.active # 获取当前活跃的sheet
    rsheet.title= "sheet1"
    rsheet['A1']= "2019.12.8-2020.1.8"
    rsheet['A2'] = "无数据"
    wbr.save("result1.xlsx")
def stage2():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['B1'] = "2020.1.8-2020.1.22"
    rsheet['B2'] = "无数据"
    wbr.save("result1.xlsx")
def stage3():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['C1'] = "2020.1.23-2020.2.7"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=2, max_row=212, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['C' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage4():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['D1'] = "2020.2.9-2020.2.13"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=213, max_row=552, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['D' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
'''
stage5：2.15  2.18  2.20  2.26  3.2
stage6：3.15  3.18  3.21  3.22  3.23  3.25
stage7：3.26  3.30  4.2  4.4  4.5  4.8  4.10
'''
def stage5():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['E1'] = "2020.2.14-2020.3.9"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=553, max_row=885, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['E' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

def stage6():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['F1'] = "2020.3.10-2020.3.25"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=886, max_row=1339, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['F' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

def stage7():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['G1'] = "2020.3.26-2020.4.15"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=1340, max_row=1655, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['G' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")


'''
stage8；4.21  4.22
stage9: 5.2  5.3  5.7  5.15
stage10: 5.18  5.19  5.20  5.23  5.24  
stage11: 6.1  
stage12：6.26  6.27
'''
def stage8():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['H1'] = "2020.4.16-2020.4.30"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=1656, max_row=1915, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['H' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage9():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['I1'] = "2020.5.1-2020.5.15"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=1916, max_row=2320, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['I' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage10():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['J1'] = "2020.5.16-2020.5.31"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=2321, max_row=2671, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['J' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage11():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['K1'] = "2020.6.1-2020.6.15"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=2672, max_row=2819, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['K' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage12():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['L1'] = "2020.6.16-2020.6.30"
    wb = load_workbook("result1.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 2
    for row in sheet.iter_rows(min_row=2820, max_row=2903, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['L' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

def stage3p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['C1'] = "2020.1.23-2020.2.7"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 213
    for row in sheet.iter_rows(min_row=1, max_row=163, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['C' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=1, max_row=46, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['C' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()

    wbr.save("result1.xlsx")
def stage4p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['D1'] = "2020.2.9-2020.2.13"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 342
    for row in sheet.iter_rows(min_row=164, max_row=409, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['D' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=47, max_row=209, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['D' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage5p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['E1'] = "2020.2.14-2020.3.9"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 335
    for row in sheet.iter_rows(min_row=410, max_row=849, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['E' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=210, max_row=1049, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['E' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage6p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['F1'] = "2020.3.10-2020.3.25"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 456
    for row in sheet.iter_rows(min_row=850, max_row=1968, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['F' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=1050, max_row=2801, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['F' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage7p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['G1'] = "2020.3.26-2020.4.15"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 318
    for row in sheet.iter_rows(min_row=1969, max_row=2496, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['G' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=2802, max_row=5142, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['G' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage8p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['H1'] = "2020.4.16-2020.4.30"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 262
    for row in sheet.iter_rows(min_row=2497, max_row=2694, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['H' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=5143, max_row=7149, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['H' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage9p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['I1'] = "2020.5.1-2020.5.15"
    wb = load_workbook("B站评论爬取3.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    i = 407
    for row in sheet.iter_rows(min_row=7150, max_row=7907, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['I' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")

def stage10p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['J1'] = "2020.5.16-2020.5.31"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 353
    for row in sheet.iter_rows(min_row=2695, max_row=2927, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['J' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=7908, max_row=8302, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['J' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage11p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['K1'] = "2020.6.1-2020.6.15"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 150
    for row in sheet.iter_rows(min_row=2928, max_row=3088, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['K' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=8303, max_row=8747, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['K' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
def stage12p():
    wbr = load_workbook('result1.xlsx')  # 实例化: 创建一个对象
    rsheet = wbr.get_sheet_by_name("sheet1")  # 获取当前活跃的sheet
    rsheet['L1'] = "2020.6.16-2020.6.30"
    wb = load_workbook("B站评论爬取2.xlsx")
    sheet = wb.get_sheet_by_name("B站评论")
    wb2 = load_workbook("B站评论爬取3.xlsx")
    sheet2 = wb2.get_sheet_by_name("B站评论")
    i = 86
    for row in sheet.iter_rows(min_row=3089, max_row=3189, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['L' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    for row in sheet2.iter_rows(min_row=8748, max_row=8836, min_col=3, max_col=3):
        for cell in row:
            if 'comments' in str(cell.value):
                pass
            elif 'None' in str(cell.value):
                pass
            elif len(str(cell.value)) == 0:
                pass
            else:
                rsheet['L' + str(i)] = cell.value
                print(cell.value, end=',')
                i += 1
        print()
    wbr.save("result1.xlsx")
main()